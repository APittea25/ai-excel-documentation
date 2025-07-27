import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from io import BytesIO
import re
import os
from collections import defaultdict
import graphviz
from docx import Document
import pandas as pd

st.set_page_config(page_title="AI-Powered Spreadsheet Documentation", layout="wide")

if "json_summaries" not in st.session_state:
    st.session_state.json_summaries = None

st.title("\U0001F4D8 AI Powered Spreadsheet Documentation")

if "expanded_all" not in st.session_state:
    st.session_state.expanded_all = False

def toggle():
    st.session_state.expanded_all = not st.session_state.expanded_all

st.button("🔁 Expand / Collapse All Named Ranges", on_click=toggle)

# Allow manual mapping of external references like [1], [2], etc.
with st.expander("🔧 Manual Mapping for External References", expanded=False):
    st.subheader("Manual Mapping for External References")
    external_refs = {}
    for i in range(1, 10):
        ref_key = f"[{i}]"
        workbook_name = st.text_input(f"Map external reference {ref_key} to workbook name (e.g., Mortality_Model_Inputs.xlsx)", key=ref_key)
        if workbook_name:
            external_refs[ref_key] = workbook_name

uploaded_files = st.file_uploader("\U0001F4C2 Upload Excel files", type=["xlsx"], accept_multiple_files=True)

if uploaded_files:  
    from file_handlers import handle_uploaded_files
    data = handle_uploaded_files(uploaded_files)
    all_named_cell_map = data["named_cell_map"]
    all_named_ref_info = data["named_ref_info"]
    file_display_names = data["file_display_names"]

    from formula_mapper import remap_formula
    
    named_ref_formulas = {}  # initialize the dictionary
    for (name, (file_name, sheet_name, coord_set, min_row, min_col)) in all_named_ref_info.items():
        entries = []
        formulas_for_graph = []

        try:
            file_bytes = file_display_names[file_name]
            wb = load_workbook(BytesIO(file_bytes.getvalue()), data_only=False)
            ws = wb[sheet_name]
            min_col_letter = get_column_letter(min([c for (_, c) in coord_set]))
            max_col_letter = get_column_letter(max([c for (_, c) in coord_set]))
            min_row_num = min([r for (r, _) in coord_set])
            max_row_num = max([r for (r, _) in coord_set])
            ref_range = f"{min_col_letter}{min_row_num}:{max_col_letter}{max_row_num}"
            cell_range = ws[ref_range] if ":" in ref_range else [[ws[ref_range]]]

            for row in cell_range:
                for cell in row:
                    row_offset = cell.row - min_row + 1
                    col_offset = cell.column - min_col + 1
                    label = f"{name}[{row_offset}][{col_offset}]"

                    try:
                        formula = None
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            formula = cell.value.strip()
                        elif hasattr(cell, 'value') and hasattr(cell.value, 'text'):
                            formula = str(cell.value.text).strip()
                        elif hasattr(cell, 'value'):
                            formula = str(cell.value)

                        if formula:
                            remapped = remap_formula(formula, file_name, sheet_name, all_named_cell_map, external_refs)
                            formulas_for_graph.append(remapped)
                        elif cell.value is not None:
                            formula = f"[value] {str(cell.value)}"
                            remapped = formula
                        else:
                            formula = "(empty)"
                            remapped = formula
                    except Exception as e:
                        formula = f"[error reading cell: {e}]"
                        remapped = formula

                    entries.append(f"{label} = {formula}\n → {remapped}")
        except Exception as e:
            entries.append(f"❌ Error accessing {name} in {sheet_name}: {e}")

        
        named_ref_formulas[name] = formulas_for_graph
        
        limit = 50
        snippet = entries if limit is None else entries[:limit]
        
        with st.expander(
            f"📌 Named Range: {name} → {sheet_name} in {file_name}",
            expanded=st.session_state.expanded_all
        ):
            st.code("\n".join(snippet), language="text")
            if limit is not None and len(entries) > limit:
                st.write(f"...and {len(entries) - limit} more lines hidden")
                
    # —– Missing direct cell references (not in any named range) —–
    with st.expander("⚠️ Missing Direct Cell References", expanded=True):
        st.markdown("#### 🔍 Check for A1-style cell references not covered by any named range")

        raw_ref_re = re.compile(r"\b([A-Z]{1,3}[0-9]{1,7})\b")
        missing_refs = defaultdict(set)

        for nm, formulas in named_ref_formulas.items():
            for f in formulas:
                for ref in raw_ref_re.findall(f):
                    if re.search(rf"\[{nm}\]\[\d+\]\[\d+\]", f):
                        continue
                    missing_refs[nm].add(ref)

        if missing_refs:
            for nm, refs in missing_refs.items():
                st.warning(
                    f"In **{nm}**, these direct cell refs weren’t wrapped by a named range: "
                    f"{', '.join(sorted(refs))}"
                )
        else:
            st.success("✅ No missing direct cell references found.")
    
    # Dependency Graph
    st.subheader("🔗 Dependency Graph")
    dot = graphviz.Digraph()
    dot.attr(compound='true', rankdir='LR')

    grouped = defaultdict(list)
    for name, (file, *_rest) in all_named_ref_info.items():
        grouped[file].append(name)

    dependencies = defaultdict(set)
    for target, formulas in named_ref_formulas.items():
        joined = " ".join(formulas)
        for source in named_ref_formulas:
            if source != target and re.search(rf"\b{re.escape(source)}\b", joined):
                dependencies[target].add(source)

    for i, (file_name, nodes) in enumerate(grouped.items()):
        with dot.subgraph(name=f"cluster_{i}") as c:
            c.attr(label=file_name)
            c.attr(style='filled', color='lightgrey')
            for node in nodes:
                c.node(node)

    for target, sources in dependencies.items():
        for source in sources:
            dot.edge(source, target)

    st.graphviz_chart(dot)

# ---- Imports for AI-Generated Response ----

    import json
    
    from prompt import (
            build_json_summary_prompt,
            build_purpose_prompt,
            build_input_prompt,
            build_output_prompt,
            build_logic_prompt,
            build_check_prompt,
            build_assumptions_prompt
        )
        
    from llm_engine import call_chat_model
    
# --- JSON Summary Generation Section ---
    
    st.subheader("🧠 Generate JSON and Documentation")

    generate_json = st.button("🧾 Generate")

    if generate_json:
        
        summaries = {}
        
        for name, formulas in named_ref_formulas.items():
            if not formulas:
                continue

            try:
                JSON_prompt = build_json_summary_prompt(name, formulas)

                response = call_chat_model(
                    system_msg="You summarize spreadsheet formulas into structured JSON.",
                    user_prompt=JSON_prompt
                )

                parsed = json.loads(response)

                # (Continue adding file_name, sheet_name, dependencies, etc.)
                file_name, sheet_name, coord_set, *_ = all_named_ref_info[name]

                min_col_letter = get_column_letter(min([c for (_, c) in coord_set]))
                max_col_letter = get_column_letter(max([c for (_, c) in coord_set]))
                min_row_num = min([r for (r, _) in coord_set])
                max_row_num = max([r for (r, _) in coord_set])
                excel_range = f"{min_col_letter}{min_row_num}:{max_col_letter}{max_row_num}"

                parsed.update({
                    "named_range": name,
                    "file_name": file_name,
                    "sheet_name": sheet_name,
                    "excel_range": excel_range,
                    "dependencies": sorted(dependencies.get(name, []))
                })

                summaries[name] = parsed

            except Exception as e:
                summaries[name] = {"named_range": name, "error": str(e)}

        with st.expander("📦 View JSON Output", expanded=False):
            st.json(summaries)
        
        #Import hints
        from hint import generate_hint_sentence
        hint_sentence = generate_hint_sentence(summaries)
        
        # --- Generate high-level Purpose description ---
        purpose_prompt = build_purpose_prompt(summaries, hint_sentence)

        model_purpose = call_chat_model(
            system_msg="You write purpose sections for actuarial models.",
            user_prompt=purpose_prompt
        )
      
        # input data
        input_summaries = {k: v for k, v in summaries.items() if k.startswith("i_")}
        inputs_data = []

        for idx, (name, summary) in enumerate(input_summaries.items(), start=1):
            excel_range = summary.get("excel_range", "")

            # Determine source based on name
            if "_a_" in name:
                source = "Assumptions team"
            elif "_m_" in name:
                source = "Modelling team"
            else:
                source = "Unknown"

            # Determine type based on excel_range
            cell_type = "Unknown"
            if ":" not in excel_range:
                cell_type = "Error"  # invalid or incomplete range
            else:
                try:
                    from_cell, to_cell = excel_range.split(":")
                    if from_cell == to_cell:
                        cell_type = "Cell"
                    else:
                        from_col = re.sub(r"\d", "", from_cell)
                        from_row = int(re.sub(r"\D", "", from_cell))
                        to_col = re.sub(r"\d", "", to_cell)
                        to_row = int(re.sub(r"\D", "", to_cell))

                        if from_col == to_col:
                            cell_type = "Vector"  # vertical
                        elif from_row == to_row:
                            cell_type = "Vector"  # horizontal
                        else:
                            cell_type = "Table"
                except Exception:
                    cell_type = "Error"

            inputs_data.append({
                "No.": idx,
                "Name": name,
                "Type": cell_type,
                "Source": source,
                "Info": ""  # To be filled by GPT
            })
            
        # GPT to populate Info field
        for row in inputs_data:
            input_name = row["Name"]
            summary_json = input_summaries[input_name]
            json_summary = summary_json.get("summary", "")
            general_formula = summary_json.get("general_formula", "")
            sheet = summary_json.get("sheet_name", "")
            excel_range = summary_json.get("excel_range", "")
            
            input_prompt = build_input_prompt(input_name, summary_json, hint_sentence)
            row["Info"] = call_chat_model(
                system_msg="You provide concise descriptions of actuarial inputs.",
                user_prompt=input_prompt
            )
        inputs_df = pd.DataFrame(inputs_data)

        # --- Output data ---
        output_summaries = {k: v for k, v in summaries.items() if k.startswith("o_")}
        outputs_data = []
        for idx, (name, summary) in enumerate(output_summaries.items(), start=1):
            outputs_data.append({
                "No.": idx,
                "Name": name,
                "Description": ""  # To be filled by GPT
            })
        output_names = " ".join(output_summaries.keys()).lower()

                   
        # GPT to populate output descriptions
        for row in outputs_data:
            output_name = row["Name"]
            summary_json = output_summaries[output_name]
            json_summary = summary_json.get("summary", "")
            general_formula = summary_json.get("general_formula", "")
            sheet = summary_json.get("sheet_name", "")
            excel_range = summary_json.get("excel_range", "")

            output_prompt = build_output_prompt(name, summary_json, hint_sentence)
            row["Description"] = call_chat_model(
                system_msg="You describe actuarial spreadsheet outputs.",
                user_prompt=output_prompt
            )
            
        outputs_df = pd.DataFrame(outputs_data)

        # --- Logic documentation based on _c1_, _c2_, etc. ---
        logic_summaries = {}
        logic_pattern = re.compile(r"^_c(\d+)_.*")  # matches _c1_, _c2_, etc.

        # Extract and sort logic blocks
        for name in summaries:
            match = logic_pattern.match(name)
            if match:
                step_number = int(match.group(1))
                logic_summaries[step_number] = name

        logic_steps = []
        for step_number in sorted(logic_summaries):
            name = logic_summaries[step_number]
            summary_json = summaries[name]
            json_summary = summary_json.get("summary", "")
            general_formula = summary_json.get("general_formula", "")
            dependencies_list = summary_json.get("dependencies", [])
            excel_range = summary_json.get("excel_range", "")
            sheet = summary_json.get("sheet_name", "")

            logic_prompt = build_logic_prompt(name, summary_json, step_number, hint_sentence)

            explanation = call_chat_model(
                system_msg="You describe logic steps in actuarial models clearly.",
                user_prompt=logic_prompt
            )
            
            logic_steps.append({
                "Step": step_number,
                "Named Range": name,
                "Description": explanation
            })

        # If no _cN_ logic blocks found, issue warning
        if not logic_steps:
            st.warning("⚠️ No logic components found using `_c1_`, `_c2_`, etc. naming convention. Please check that named ranges follow this format.")

        # --- Checks and Validation ---
        check_pattern = re.compile(r"^_ch(\d+)_.*")  # matches _ch1_, _ch2_, etc.
        check_summaries = {}

        # Extract and order checks
        for name in summaries:
            match = check_pattern.match(name)
            if match:
                check_num = int(match.group(1))
                check_summaries[check_num] = name

        check_data = []
        for check_num in sorted(check_summaries):
            name = check_summaries[check_num]
            summary_json = summaries[name]
            json_summary = summary_json.get("summary", "")
            general_formula = summary_json.get("general_formula", "")
            sheet = summary_json.get("sheet_name", "")
            excel_range = summary_json.get("excel_range", "")

            # GPT prompt to describe what the check does
            check_prompt = build_check_prompt(name, summary_json, hint_sentence)

            description = call_chat_model(
                system_msg="You describe spreadsheet checks in actuarial models.",
                user_prompt=check_prompt
            )
            
            check_data.append({
                "Check No.": check_num,
                "Named Range": name,
                "Description": description
            })

        # Convert to DataFrame for Streamlit and Word doc
        checks_df = pd.DataFrame(check_data)

        # --- Assumptions and Limitations ---
        try:
            all_summaries = "\n".join(
                f"{k}: {v.get('summary', '')}" for k, v in summaries.items() if "summary" in v
            )
            all_formulas = "\n".join(
                f"{k}: {v.get('general_formula', '')}" for k, v in summaries.items() if "general_formula" in v
            )

            assumptions_prompt = build_assumptions_prompt(summaries, hint_sentence)

            assumptions_text = call_chat_model(
                system_msg="You describe assumptions and limitations in actuarial spreadsheet models.",
                user_prompt=assumptions_prompt
            )

        except Exception as e:
            assumptions_text = f"Error generating assumptions and limitations: {e}"
        
        
        with st.expander("📄 Spreadsheet Document", expanded=False):
            st.title("📄 Model Documentation")
            st.header("## Version Control")
            # Model version control table
            st.subheader("### Model Version Control")
            model_version_df = pd.DataFrame({
                "Version": [""],
                "Date": ["00/00/0000"],
                "Info": [""],
                "Updated by": [""],
                "Reviewed by": [""],
                "Review Date": [""]
            })
            st.dataframe(model_version_df, use_container_width=True)

            # Documentation version control table
            st.subheader("### Documentation Version Control")
            doc_version_df = pd.DataFrame({
                "Version": [""],
                "Date": ["00/00/0000"],
                "Info": [""],
                "Updated by": [""],
                "Reviewed by": [""],
                "Review Date": [""]
            })
            st.dataframe(doc_version_df, use_container_width=True)

            # Ownership section
            st.header("## Ownership")
            st.text("### Owner")
            st.text("### Risk rating (or other client control standard)")
            st.text("### Internal audit history")

            # Purpose
            st.header("## Purpose")
            st.text_area("Describe the purpose of the model:", value=model_purpose)

            # Inputs table
            st.header("## Inputs")
            row_height = 35
            max_height = 500
            calculated_height = min(len(inputs_df) * row_height + 35, max_height)
            st.dataframe(inputs_df, use_container_width=True)

            # Outputs
            st.header("## Outputs")
            st.dataframe(outputs_df, use_container_width=True)

            # Logic
            st.header("## Logic")
            if logic_steps:
                logic_df = pd.DataFrame(logic_steps)
                st.dataframe(logic_df, use_container_width=True)
            else:
                st.warning("⚠️ No logic documented. Ensure named ranges follow `_c1_`, `_c2_`, ... convention.")

            # Checks and validation
            st.header("## Checks and Validation")
            if not checks_df.empty:
                st.dataframe(checks_df, use_container_width=True)
            else:
                st.warning("⚠️ No checks found using `_ch1_`, `_ch2_`, etc. naming convention. The model may not include validation steps.")

            # Assumptions and limitations
            st.header("## Assumptions and Limitations")
            st.text_area("List assumptions and limitations:", value=assumptions_text, height=200)

            # TAS Compliance
            st.header("## TAS Compliance")
            st.text_area("Describe how the model complies with TAS:")

        # JSON download
        json_str = json.dumps(summaries, indent=2)
        st.download_button("📥 Download JSON Summary", data=json_str, file_name="named_range_summaries.json", mime="application/json")

        from doc_builder import build_word_doc
        docx_io = build_word_doc(
            summaries=summaries,
            model_purpose=model_purpose,
            inputs_data=inputs_data,
            outputs_data=outputs_data,
            logic_steps=logic_steps,
            checks_data=check_data,
            assumptions_text=assumptions_text
        )
        
        st.download_button(
            "📄 Download Summary as Word Document",
            data=docx_io,
            file_name="named_range_summary.docx",
            mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        )

    else:
        st.info("Press the button above to generate a GPT-based JSON and Documentation.")

else:
    st.info("⬆️ Upload one or more .xlsx files to begin.")
