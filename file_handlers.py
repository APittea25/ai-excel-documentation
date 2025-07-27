# file_handlers.py
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from io import BytesIO

def handle_uploaded_files(uploaded_files):
    all_named_cell_map = {}
    all_named_ref_info = {}
    file_display_names = {}
    named_ref_formulas = {}

    for uploaded_file in uploaded_files:
        display_name = uploaded_file.name
        file_display_names[display_name] = uploaded_file
        wb = load_workbook(BytesIO(uploaded_file.read()), data_only=False)

        for name in wb.defined_names:
            dn = wb.defined_names[name]
            if dn.is_external or not dn.attr_text:
                continue
            for sheet_name, ref in dn.destinations:
                try:
                    ws = wb[sheet_name]
                    ref_clean = ref.replace("$", "").split("!")[-1]
                    cells = ws[ref_clean] if ":" in ref_clean else [[ws[ref_clean]]]

                    min_row = min(cell.row for row in cells for cell in row)
                    min_col = min(cell.column for row in cells for cell in row)

                    coord_set = set()
                    formulas_for_graph = []
                    for row in cells:
                        for cell in row:
                            r, c = cell.row, cell.column
                            row_offset = r - min_row + 1
                            col_offset = c - min_col + 1
                            all_named_cell_map[(display_name, sheet_name, r, c)] = (name, row_offset, col_offset)
                            coord_set.add((r, c))
                            if isinstance(cell.value, str) and cell.value.startswith("="):
                                formulas_for_graph.append(cell.value.strip())
                            elif cell.value is not None:
                                formulas_for_graph.append(str(cell.value))
                    all_named_ref_info[name] = (display_name, sheet_name, coord_set, min_row, min_col)
                    named_ref_formulas[name] = formulas_for_graph
                except Exception:
                    continue

    return {
        "named_cell_map": all_named_cell_map,
        "named_ref_info": all_named_ref_info,
        "file_display_names": file_display_names,
        "named_ref_formulas": named_ref_formulas 
    }
