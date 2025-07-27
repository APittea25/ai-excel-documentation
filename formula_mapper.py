# formula_mapper.py
import re
from openpyxl.utils import column_index_from_string, get_column_letter

def remap_formula(formula, current_file, current_sheet, all_named_cell_map, external_refs):
    if not formula:
        return ""

    def cell_address(row, col):
        return f"{get_column_letter(col)}{row}"

    def remap_single_cell(ref, default_file, default_sheet):
        match = re.match(r"(?:'([^']+)'|([^'!]+))!([$A-Z]+[0-9]+)", ref)
        if match:
            sheet_name = match.group(1) or match.group(2)
            addr = match.group(3)
            external_match = re.match(r"\[(\d+)\]", sheet_name)
            if external_match:
                external_ref = external_match.group(0)
                external_file = external_refs.get(external_ref, external_ref)
                return f"[{external_file}]{ref}"
        else:
            sheet_name = default_sheet
            addr = ref

        addr = addr.replace("$", "").upper()
        match = re.match(r"([A-Z]+)([0-9]+)", addr)
        if not match:
            return ref
        col_str, row_str = match.groups()
        row = int(row_str)
        col = column_index_from_string(col_str)

        key = (default_file, sheet_name, row, col)
        if key in all_named_cell_map:
            name, r_off, c_off = all_named_cell_map[key]
            return f"[{default_file}]{name}[{r_off}][{c_off}]"
        else:
            return f"{sheet_name}!{addr}"

    def remap_range(ref, default_file, default_sheet):
        if ref.startswith("["):
            match = re.match(r"\[(\d+)\]", ref)
            if match:
                external_ref = match.group(0)
                external_file = external_refs.get(external_ref, external_ref)
                return f"[{external_file}]{ref}"

        match = re.match(r"(?:'([^']+)'|([^'!]+))!([$A-Z]+[0-9]+(?::[$A-Z]+[0-9]+)?)", ref)
        if match:
            sheet_name = match.group(1) or match.group(2)
            addr = match.group(3)
        else:
            sheet_name = default_sheet
            addr = ref

        addr = addr.replace("$", "").upper()
        if ":" not in addr:
            return remap_single_cell(ref, default_file, default_sheet)

        start, end = addr.split(":")
        m1 = re.match(r"([A-Z]+)([0-9]+)", start)
        m2 = re.match(r"([A-Z]+)([0-9]+)", end)
        if not m1 or not m2:
            return ref
        start_col = column_index_from_string(m1.group(1))
        start_row = int(m1.group(2))
        end_col = column_index_from_string(m2.group(1))
        end_row = int(m2.group(2))

        label_set = set()
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                key = (default_file, sheet_name, row, col)
                if key in all_named_cell_map:
                    name, r_off, c_off = all_named_cell_map[key]
                    label_set.add(f"[{default_file}]{name}[{r_off}][{c_off}]")
                else:
                    label_set.add(f"{sheet_name}!{get_column_letter(col)}{row}")
        return ", ".join(sorted(label_set))

    pattern = r"(?<![A-Za-z0-9_])(?:'[^']+'|[A-Za-z0-9_]+)!\$?[A-Z]{1,3}\$?[0-9]{1,7}(?::\$?[A-Z]{1,3}\$?[0-9]{1,7})?|(?<![A-Za-z0-9_])\$?[A-Z]{1,3}\$?[0-9]{1,7}(?::\$?[A-Z]{1,3}\$?[0-9]{1,7})?"
    matches = list(re.finditer(pattern, formula))
    replaced_formula = formula
    offset = 0
    for match in matches:
        raw = match.group(0)
        remapped = remap_range(raw, current_file, current_sheet)
        start, end = match.start() + offset, match.end() + offset
        replaced_formula = replaced_formula[:start] + remapped + replaced_formula[end:]
        offset += len(remapped) - len(raw)

    return replaced_formula
